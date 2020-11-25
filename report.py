import pandas
from collections import defaultdict, Counter
from openpyxl import load_workbook
# Импорт необходимых библиотек и модулей

#Читаем файл с логами и создаем словарь
excel_data = pandas.read_excel('logs.xlsx', sheet_name='log')
excel_data_dict = excel_data.to_dict(orient='records')

#создаем списки где будут храниться данные
browsers = []
goods = []
man_goods = []
woman_goods = []


#Заполняем списки
for element in excel_data_dict:
    browsers.append(element['Браузер'])
    good = element['Купленные товары'].split(',') 
    for g in good:
        goods.append(g.strip()) 
        if element['Пол']=='м':
            man_goods.append(g.strip()) 
        else:
            woman_goods.append(g.strip()) 


def popular(listing, keys, top):
    popular_listing = Counter(listing).most_common(top) # список популярных объектов
    popular_list=[]  
    for element in popular_listing:
        m=[0]*13 # m= month M[0]  общее число покупок m[1]- m[12] по каждому месяцу
        m[0]=element[1] #присваиваем значение общих покупок
        for i in range(len(excel_data_dict)):   #проходимся по всему логу и останавливаемся на ключе который мы задали
            if element[0] in excel_data_dict[i][keys]:  #если мы его нашли,то дальше проходмся по логу
                date = excel_data_dict[i]['Дата посещения']  #находим дату из нее извлекаем месяц
                date_str = str(date)
                date_notime = date_str.split()[0]
                month = date_notime.split('-')[1] #извлекли месяц
                m[int(month)] +=1 #добавляем кол-во покупок к каждому месяцу
        popular_list.append([element[0], m[0], m[1], m[2], m[3], m[4], m[5], m[6], m[7], m[8], m[9], m[10], m[11], m[12]])  #формируем список
    return popular_list

#находим необходмые значения для таблицы
browser_list = popular(browsers, 'Браузер', 7) 
goods_list = popular(goods, 'Купленные товары', 7) 
goods_man = Counter(man_goods).most_common() 
popular_goods_man = goods_man[0][0] 
nopopular_goods_man = goods_man[-1][0] 
goods_woman = Counter(woman_goods).most_common() 
popular_goods_woman = goods_woman[0][0] 
nopopular_goods_woman = goods_woman[-1][0]  


#Запись полученных данных в excel
wb = load_workbook(filename='report.xlsx')
sheet = wb['Лист1']
for i in range(7):
    for j in range(14):
        sheet.cell(row=i+5, column=j+1).value = browser_list[i][j]
        sheet.cell(row=i+19, column=j+1).value = goods_list[i][j]
sheet["B31"] = popular_goods_man
sheet["B32"] = popular_goods_woman
sheet["B33"] = nopopular_goods_man
sheet["B34"] = nopopular_goods_woman
wb.save(filename='report.xlsx')